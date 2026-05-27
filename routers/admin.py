import hashlib
import logging
from collections import Counter
from io import BytesIO
from datetime import datetime

import openpyxl
from fastapi import APIRouter, BackgroundTasks, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse

from core.template import templates
from core.config import MAX_UPLOAD_MB
from services.document_service import insert_document_chunks
from services.embedding_service import index_document_chunks
from services.auth_service import is_admin
from services.file_service import extract_text_from_file
from repositories.group_repository import get_groups, insert_group, delete_group_if_empty
from repositories.document_repository import get_documents, insert_document, delete_document_and_chunks, document_exists_by_hash
from repositories.vector_repository import delete_embeddings_by_document
from repositories.user_repository import get_all_users, create_user, delete_user, username_exists, get_user_by_id
from repositories.user_group_repository import get_user_group_ids, set_user_groups, get_users_with_group_counts
from repositories.chat_history_repository import (
    get_all_daily_query_counts, get_hourly_query_counts,
    get_raw_group_ids_all
)
from services.auth_service import hash_password

logger = logging.getLogger(__name__)
router = APIRouter()


def _require_admin(request: Request):
    """Return username if admin, else None."""
    username = request.session.get("username")
    role = request.session.get("role")
    if not username or role != "admin":
        return None
    return username


@router.get("/admin/dashboard", response_class=HTMLResponse)
def admin_dashboard(request: Request):
    username = _require_admin(request)
    if not username:
        return RedirectResponse(url="/", status_code=303)
    return templates.TemplateResponse(
        request=request,
        name="admin/admin_dashboard.html",
        context={"username": username, "message": "Admin 管理目錄"}
    )


@router.get("/admin/upload", response_class=HTMLResponse)
def admin_upload_page(request: Request):
    username = _require_admin(request)
    if not username:
        return RedirectResponse(url="/", status_code=303)
    return templates.TemplateResponse(
        request=request,
        name="admin/admin_upload.html",
        context={
            "username": username,
            "groups": get_groups(),
            "documents": get_documents(),
            "message": "上傳文件"
        }
    )


@router.get("/admin/groups", response_class=HTMLResponse)
def admin_groups_page(request: Request):
    username = _require_admin(request)
    if not username:
        return RedirectResponse(url="/", status_code=303)
    return templates.TemplateResponse(
        request=request,
        name="admin/admin_groups.html",
        context={"username": username, "groups": get_groups(), "message": "新建群組"}
    )


@router.post("/add-group", response_class=HTMLResponse)
def add_group(
    request: Request,
    group_name: str = Form(...),
    group_description: str = Form(...)
):
    username = _require_admin(request)
    if not username:
        return RedirectResponse(url="/", status_code=303)

    insert_group(group_name, group_description)

    return templates.TemplateResponse(
        request=request,
        name="admin/admin_groups.html",
        context={
            "message": f"群組新增成功：{group_name}",
            "username": username,
            "groups": get_groups()
        }
    )


@router.post("/upload-document", response_class=HTMLResponse)
async def upload_document(
    request: Request,
    background_tasks: BackgroundTasks,
    title: str = Form(...),
    group_id: int = Form(...),
    files: list[UploadFile] = File(...)
):
    username = _require_admin(request)
    if not username:
        return RedirectResponse(url="/", status_code=303)

    success_files = []
    failed_files = []

    for file in files:
        content_bytes = await file.read()

        if len(content_bytes) > MAX_UPLOAD_MB * 1024 * 1024:
            logger.warning("upload reject: file=%s size=%dMB exceeds limit %dMB",
                           file.filename, len(content_bytes) // (1024 * 1024), MAX_UPLOAD_MB)
            failed_files.append(f"{file.filename}（檔案過大，上限 {MAX_UPLOAD_MB}MB）")
            continue

        file_hash = hashlib.sha256(content_bytes).hexdigest()

        if document_exists_by_hash(file_hash):
            logger.info("upload skip: file=%s already exists (hash match)", file.filename)
            failed_files.append(f"{file.filename}（已上傳過）")
            continue

        content_text = extract_text_from_file(file.filename, content_bytes)

        if content_text is None:
            failed_files.append(f"{file.filename}（不支援的格式）")
            continue

        if not content_text.strip():
            failed_files.append(f"{file.filename}（沒有擷取到文字）")
            continue

        final_title = title.strip() if len(files) == 1 and title.strip() else file.filename

        document_id = insert_document(
            final_title, file.filename, content_text, group_id, file_hash
        )
        insert_document_chunks(document_id, content_text)
        background_tasks.add_task(index_document_chunks, document_id)
        success_files.append(file.filename)
        logger.info("upload success: user=%s file=%s document_id=%d group_id=%d",
                    username, file.filename, document_id, group_id)

    if success_files and failed_files:
        message = f"成功 {len(success_files)} 個，失敗 {len(failed_files)} 個"
    elif success_files:
        message = f"成功上傳 {len(success_files)} 個檔案（向量索引建立中，稍後即可搜尋）"
    else:
        message = "上傳失敗，可能有以下原因：\n" + "\n".join(failed_files)

    return templates.TemplateResponse(
        request=request,
        name="admin/admin_upload.html",
        context={
            "message": message,
            "username": username,
            "groups": get_groups(),
            "documents": get_documents()
        }
    )


@router.get("/admin/delete", response_class=HTMLResponse)
def admin_delete_page(request: Request, group_id: int = None):
    username = _require_admin(request)
    if not username:
        return RedirectResponse(url="/", status_code=303)

    documents = get_documents(group_id) if group_id else []

    return templates.TemplateResponse(
        request=request,
        name="admin/admin_delete.html",
        context={
            "username": username,
            "groups": get_groups(),
            "documents": documents,
            "selected_group_id": group_id,
            "message": "刪除文件"
        }
    )


@router.post("/delete-document")
def delete_document(request: Request, document_id: int = Form(...), group_id: int = Form(...)):
    username = _require_admin(request)
    if not username:
        return RedirectResponse(url="/", status_code=303)

    delete_document_and_chunks(document_id)
    delete_embeddings_by_document(document_id)
    return RedirectResponse(url=f"/admin/delete?group_id={group_id}", status_code=303)


@router.get("/admin/users", response_class=HTMLResponse)
def admin_users_page(request: Request):
    username = _require_admin(request)
    if not username:
        return RedirectResponse(url="/", status_code=303)
    return templates.TemplateResponse(
        request=request,
        name="admin/admin_users.html",
        context={
            "username": username,
            "users": get_all_users(),
            "group_counts": get_users_with_group_counts(),
            "message": "帳號管理",
            "error": None,
            "success": None,
        }
    )


@router.post("/admin/create-user", response_class=HTMLResponse)
def create_user_route(
    request: Request,
    new_username: str = Form(...),
    new_password: str = Form(...),
    role: str = Form(...)
):
    username = _require_admin(request)
    if not username:
        return RedirectResponse(url="/", status_code=303)

    error = None
    success = None

    if role not in ("admin", "user"):
        error = "角色必須是 admin 或 user。"
    elif not new_username.strip():
        error = "帳號名稱不能為空白。"
    elif not new_password:
        error = "密碼不能為空白。"
    elif username_exists(new_username.strip()):
        error = f"帳號「{new_username.strip()}」已存在。"
    else:
        create_user(new_username.strip(), hash_password(new_password), role)
        logger.info("admin=%s created user=%s role=%s", username, new_username.strip(), role)
        success = f"帳號「{new_username.strip()}」建立成功，角色：{role}。"

    return templates.TemplateResponse(
        request=request,
        name="admin/admin_users.html",
        context={
            "username": username,
            "users": get_all_users(),
            "group_counts": get_users_with_group_counts(),
            "message": "帳號管理",
            "error": error,
            "success": success,
        }
    )


@router.post("/admin/delete-user")
def delete_user_route(request: Request, user_id: int = Form(...)):
    username = _require_admin(request)
    if not username:
        return RedirectResponse(url="/", status_code=303)
    logger.info("admin=%s deleted user_id=%d", username, user_id)
    delete_user(user_id)
    return RedirectResponse(url="/admin/users", status_code=303)


@router.post("/delete-group", response_class=HTMLResponse)
def delete_group(request: Request, group_id: int = Form(...)):
    username = _require_admin(request)
    if not username:
        return RedirectResponse(url="/", status_code=303)

    success = delete_group_if_empty(group_id)
    message = (
        f"群組刪除成功，ID = {group_id}"
        if success else "此群組底下仍有文件，請先刪除文件後再刪除群組"
    )

    return templates.TemplateResponse(
        request=request,
        name="admin/admin_groups.html",
        context={"message": message, "username": username, "groups": get_groups()}
    )


@router.get("/admin/user-groups/{user_id}", response_class=HTMLResponse)
def admin_user_groups_page(request: Request, user_id: int):
    username = _require_admin(request)
    if not username:
        return RedirectResponse(url="/", status_code=303)

    target_user = get_user_by_id(user_id)
    if not target_user:
        return RedirectResponse(url="/admin/users", status_code=303)

    assigned_ids = set(get_user_group_ids(user_id))
    all_groups = get_groups()

    return templates.TemplateResponse(
        request=request,
        name="admin/admin_user_groups.html",
        context={
            "username": username,
            "target_user": target_user,
            "groups": all_groups,
            "assigned_ids": assigned_ids,
            "message": f"管理 {target_user.username} 的群組權限",
        }
    )


@router.post("/admin/user-groups/{user_id}", response_class=HTMLResponse)
def admin_user_groups_save(
    request: Request,
    user_id: int,
    group_ids: list[int] = Form(default=[]),
):
    username = _require_admin(request)
    if not username:
        return RedirectResponse(url="/", status_code=303)

    target_user = get_user_by_id(user_id)
    if not target_user:
        return RedirectResponse(url="/admin/users", status_code=303)

    set_user_groups(user_id, group_ids)
    logger.info("admin=%s set groups=%s for user_id=%d", username, group_ids, user_id)

    assigned_ids = set(group_ids)
    return templates.TemplateResponse(
        request=request,
        name="admin/admin_user_groups.html",
        context={
            "username": username,
            "target_user": target_user,
            "groups": get_groups(),
            "assigned_ids": assigned_ids,
            "message": f"{target_user.username} 的群組權限已更新",
            "success": True,
        }
    )


def _build_stats_data() -> dict:
    all_groups = get_groups()
    group_map = {g["id"]: g["name"] for g in all_groups}

    raw_group_ids = get_raw_group_ids_all()
    group_query_counter = Counter()
    for gids_str in raw_group_ids:
        for gid in gids_str.split(","):
            gid = gid.strip()
            if gid.isdigit():
                group_query_counter[int(gid)] += 1

    group_stats = [
        {"name": group_map.get(gid, f"群組 {gid}"), "count": cnt}
        for gid, cnt in sorted(group_query_counter.items(), key=lambda x: -x[1])
    ]

    daily_counts = []
    try:
        daily_counts = get_all_daily_query_counts()
    except Exception as e:
        logger.warning("get_all_daily_query_counts failed: %s", e)

    hourly_data = []
    try:
        hourly_data = get_hourly_query_counts()
    except Exception as e:
        logger.warning("get_hourly_query_counts failed: %s", e)

    return {"group_stats": group_stats, "daily_counts": daily_counts, "hourly_data": hourly_data}


@router.get("/admin/stats", response_class=HTMLResponse)
def admin_stats_page(request: Request):
    username = _require_admin(request)
    if not username:
        return RedirectResponse(url="/", status_code=303)

    data = _build_stats_data()
    return templates.TemplateResponse(
        request=request,
        name="admin/admin_stats.html",
        context={"username": username, "message": "使用統計", **data}
    )


@router.get("/admin/stats/export")
def admin_stats_export(request: Request):
    username = _require_admin(request)
    if not username:
        return RedirectResponse(url="/", status_code=303)

    data = _build_stats_data()

    wb = openpyxl.Workbook()

    # Sheet 1: 每日問答量
    ws1 = wb.active
    ws1.title = "每日問答量"
    ws1.append(["日期", "問答次數"])
    for row in data["daily_counts"]:
        ws1.append([row["day"], row["count"]])

    # Sheet 2: 各群組查詢次數
    ws2 = wb.create_sheet("各群組查詢次數")
    ws2.append(["群組名稱", "查詢次數"])
    for row in data["group_stats"]:
        ws2.append([row["name"], row["count"]])

    # Sheet 3: 使用時段分析
    ws3 = wb.create_sheet("使用時段分析")
    ws3.append(["時段", "查詢次數"])
    for row in data["hourly_data"]:
        ws3.append([f"{row['hour']:02d}:00", row["count"]])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
